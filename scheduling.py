import openpyxl
import random

from techns import Techn
from problem import Problem

class Scheduling(Techn,Problem):
    schedule = list()
    techns_ids = list()
    times = list()

    def return_problem_duration(self, name_pm):
        wb = openpyxl.load_workbook('problems_types.xlsx')
        tp_sheet = wb['types']
        rows = tp_sheet.max_row
        for i in range(2, rows + 1):
            if (tp_sheet.cell(row=i, column=2).value == name_pm):
                return tp_sheet.cell(row=i, column=3).value

    def list_of_lists(self):
        clientfile = openpyxl.load_workbook('excel_files\clients.xlsx')
        clientsheet = clientfile['clients']
        maxrw = clientsheet.max_row  # maxrow is the number of tech
        listoflists = []
        for i in range(2, maxrw + 10):
            sublist = [0, 0]
            listoflists.append(sublist)
        return listoflists

    def init_empty_schedule(self, length_of_schedule):
        self.schedule = [[] for i in range(length_of_schedule)]
        return

    def init_times(self, length_of_schedule):
        self.times = [60*8 for i in range(length_of_schedule)]
        return

    def init_techn_ids(self):
        t=Techn()
        t.get_all_ids()
        techns_ids =t.techID

    def get_problems_id_time(self,type_of_problem):
        p = Problem()
        all_problems = p.get_list_id_time(type_of_problem)
        return all_problems

    def update_time(self, current_techn, time):
        self.times[current_techn] -= time

    def problems_schedule(self, type_of_problem):
        all_problems = self.get_problems_id_time(type_of_problem)
        num_of_techns=len(self.techns_ids)
        current_techn = 0
        for problem in all_problems:
            test = 0
            has_time = True  # hasTime(current_techn, problem[1])
            while not has_time :
                current_techn = (current_techn + 1)%num_of_techns
                test += 1
                if test == num_of_techns:
                    break
                has_time = True  # hasTime(current_techn, problem[1])
            if has_time:
                self.schedule[current_techn].append(problem)
                self.update_time(current_techn, problem[1])
                current_techn = (current_techn + 1) % num_of_techns
        # save_schedule()
        # report_errors()

    def main_schedule(self):
        #initialize global vars
        self.init_techn_ids()
        self.init_empty_schedule(len(self.techns_ids))
        self.init_times(len(self.techns_ids))

        self.problems_schedule('critical')
        self.problems_schedule('regular')
