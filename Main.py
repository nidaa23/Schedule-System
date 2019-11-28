import sys
import openpyxl
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5 import  QtGui,QtCore
from PyQt5.QtCore import *
from login_forum import Ui_LoginWindow
from reportG import Ui_Dialog_Client
from clientCls import Client
from OrganizedGUI import Ui_Dialog_Admin
from PyQt5.QtWidgets import QApplication
import PyQt5.uic as  uic
import os
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5 import  QtGui,QtCore
from PyQt5.QtCore import *
from login_forum import Ui_LoginWindow
from report_problem import Problem
from product import Product
from techns import Techn



class main_login(QMainWindow,Ui_LoginWindow,Problem,Ui_Dialog_Client,Ui_Dialog_Admin,Techn,Product,Client):

    def __init__(self):
        Problem.__init__(self)
        Product.__init__(self)
        Client.__init__(self)
        Techn.__init__(self)
        QMainWindow.__init__(self)
        Ui_LoginWindow.__init__(self)
        self.setupUi(self)
        self.btn_login.clicked.connect(self.redirectTo)


    def redirectTo(self):
        username_ = self.user_name.text()
        pwd = self.user_pass.text()
        res_client = self.client_login(username_, pwd)
        res_tech = self.techn_login(username_, pwd)
        res_admin = self.admin_login(username_,pwd)


        if(res_client == 1):

            self.clientWindow = QtWidgets.QMainWindow()
            self.id_client = self.user_pass.text()
            self.ui = Ui_Dialog_Client(self.id_client)
            self.ui.setupUi(self.clientWindow)
            self.hide()
            self.clientWindow.show()

            self.ui.report_p()

            self.ui.send.clicked.connect(self.send_data)

        elif (res_admin == 1):
            QMainWindow.__init__(self)
            Ui_Dialog_Admin.__init__(self)

            self.adminWindow = QtWidgets.QMainWindow()
            self.ui = Ui_Dialog_Admin()
            self.ui.setupUi_(self.adminWindow)
            self.hide()
            self.adminWindow.show()
            self.off()
            self.ui.add_clientB.clicked.connect(self.addclient)
            self.ui.add_productB.clicked.connect(self.addproduct)
            self.ui.add_techB.clicked.connect(self.addtech)
            self.ui.push_product.clicked.connect(self.pushproduct)
            self.ui.push_client.clicked.connect(self.pushClient)
            self.ui.push_tech.clicked.connect(self.pushtech)






    # Authentication methods
    def client_login(self,name,pwdId):
        clientfile = openpyxl.load_workbook('excel_files\\clients.xlsx')
        clientsheet = clientfile['clients']
        row_s = clientsheet.max_row
        for i in range(2, row_s + 1):
            if (str(clientsheet.cell(row=i, column=2).value) == str(pwdId) and clientsheet.cell(row=i, column=1).value == name):
                return 1
        return 0

    def techn_login(self,username, pwd):
        tech_file = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = tech_file['techns']
        rows = tc_sheet.max_row
        for i in range(2, rows + 1):
            if (str(tc_sheet.cell(row=i, column=3).value) == str(username) and  str(tc_sheet.cell(row=i,column=4).value) == str(pwd)):
                return 1
        return 0

    def admin_login(self, username,pwd):
        if(username=="admin" and pwd=="0123456"):
            return 1
        return 0
#Client page methods:
    def report_p(self):
        id_client=self.user_pass.text()
        proud_lst=list(self.get_client_product())
        prob_lst=list(self.get_problems_types())
        for i in proud_lst:
            self.list_product.addItem(i[1])
        for j in prob_lst:
            self.lis_problem.addItem(j)
        self.ui.groupBox.show()
        self.list_product.currentTextChanged.connect(self.select_product)
        self.lis_problem.currentTextChanged.connect(self.select_problem)


    def select_product(self):
        choice_product=self.list_product.currentText()
        return choice_product
       # choice_product=self.list_product.currentText()
    def select_problem(self):
        choice_problem=self.lis_problem.currentText()
        return choice_problem
    def send_data(self):
        id_client = self.id_client
        ch_product=self.ui.select_product()
        ch_problem=self.ui.select_problem()
        self.add_problem_to_file(ch_product,id_client,ch_problem)
        print(id_client,ch_problem,ch_product)
        #self.add_problem_to_file(choice_product,id_client,choice_problem)
    def get_problems_types(self):
        lsproblems=[]
        f=openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        sheet1 = f['types']
        row = sheet1.max_row + 1
        for i in range(2,row):
            lsproblems.append(sheet1.cell(row=i, column=2).value)
        return lsproblems

    def get_client_product(self):
        ans = []
        wb = openpyxl.load_workbook('excel_files\\products.xlsx')
        sheet1 = wb['products']
        for i in range(2, sheet1.max_column + 1):
            if (str(sheet1.cell(row=i, column=3).value) == str(self.id_client)):
                ans.append((sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value))
        return ans

    #Admin page methods:
    def off(self):
        self.ui.groupBox1.hide()
        self.ui.groupBox2.hide()
        self.ui.groupBox3.hide()

    def addclient(self):
        self.off()
        self.ui.groupBox1.show()

    def addproduct(self):
        self.off()
        self.ui.groupBox2.show()
        self.ui.groupBox2.setGeometry(400, 20, 281, 221)

    def addtech(self):
        self.off()
        self.ui.groupBox3.show()
        self.ui.groupBox3.setGeometry(400, 20, 281, 221)

    def pushproduct(self):
        name = self.ui.product_name.text()
        id_client = self.ui.id_pclient.text()
        self.add_product(name, id_client)

    def pushClient(self):
        n = self.ui.client_name.text()
        lo = self.ui.client_location.text()
        print(n, lo)
        self.add_client_to_excel(n,lo)
        # self.add_client_to_excel(name,loc)
        # self.add_client_to_excel(name,loc)

    def pushtech(self):
        name = self.ui.tech_name.text()
        username = self.ui.tech_username.text()
        password = self.ui.tech_pass.text()
        self.add_tech(name, username, password)








if __name__=="__main__":
    app=QApplication(sys.argv)
    window=main_login()
    window.show()
    sys.exit(app.exec_())
